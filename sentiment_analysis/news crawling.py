import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import datetime as dt
import re
import random
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 언론사별 상세 페이지 본문 CSS 셀렉터
SITE_BODY_SELECTORS = {
    'chosun.com':            ['section.article-body p.article-body_content'],
    'news-today.co.kr':      ['div.view_con_wrap p[style*="text-align"]'],
    'financetoday.co.kr':    ['div#article-view-content-div > p'],
    'asiatoday.co.kr':       ['article#article-view-content-div > p'],
    'etoday.co.kr':          ['div.articleView[itemprop="articleBody"] > p'],
    'newdaily.co.kr':        ['div#article_conent[itemprop="articleBody"] li.par > div'],
    'edaily.co.kr':          ['div.news_body'],
    'hankyung.com':          ['div#articletxt.article-body'],
    'sedaily.com':           ['div.article_view[itemprop="articleBody"]'],
    'mk.co.kr':              ['div.news_cnt_detail_wrap[itemprop="articleBody"] > p[refid]'],
    'wideopinion.co.kr':     ['article#article-view-content-div > p'],
    'thepublic.kr':          ['article#article-view-content-div > p'],
    'asiae.co.kr':           ['div#txt_area > p[data-alda-marking]'],
    'heraldcorp.com':        ['article#dic_area.go_trans._article_content'],
    'fnnews.com':            ['div#article_content.cont_view'],
    'yna.co.kr':             ['div#articleBody[itemprop="articleBody"] > div.article-body-text'],
    'cbs.co.kr':             ['div#CmAdContent > p'],
    'hankookilbo.com':       ['div.col-main[itemprop="articleBody"] > p.editor-p.read[data-break-type="text"]'],
    'donga.com':             ['section.news_view > div.view_m_adK'],
    'joongang.co.kr':        ['div#article_body p'],
    'hani.co.kr':            ['div.article-text > p.text'],
    'newsis.com':            ['div.thumCont'],
    'kbs.co.kr':             ['div#cont_newstext.detail-body'],
    'sbs.co.kr':             ['div#cnbc-front-articleContent-area-font'],
    'kyunghyang.com':        ['div#articleBody.art_body > p.content_text'],
    'thefact.co.kr':         ['div#content_area.atcRead > p'],
    # 필요시 추가
}

def clean_text(text: str) -> str:
    text = re.sub(r'[^A-Za-z0-9\uAC00-\uD7A3.%]', ' ', text)
    return re.sub(r' +', ' ', text).strip()

def get_article_content(url: str) -> str:
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
        "Accept-Language": "ko-KR,ko;q=0.9",
    }
    try:
        resp = requests.get(url, headers=headers, timeout=10)
        resp.encoding = resp.apparent_encoding
        if resp.status_code != 200:
            print(f"[!] 응답 실패: {resp.status_code} – {url}")
            return ""

        soup = BeautifulSoup(resp.text, 'html.parser')
        for tag in soup(['script', 'style', 'aside', 'iframe']):
            tag.decompose()

        domain = re.sub(r'^www\.', '', re.sub(r'^https?://', '', url)).split('/')[0]
        selectors = SITE_BODY_SELECTORS.get(domain, ['div[itemprop="articleBody"]'])

        for sel in selectors:
            node = soup.select_one(sel)
            if not node:
                continue

            # 1) <p> 태그 우선 추출
            ps = node.find_all('p')
            if ps:
                txt = ' '.join(clean_text(p.get_text()) for p in ps if p.get_text())
                if len(txt) > 50:
                    return txt

            # 2) <br> 태그 기준 분리
            html = node.decode_contents()
            parts = [
                clean_text(BeautifulSoup(seg, 'html.parser').get_text())
                for seg in html.split('<br') if seg.strip()
            ]
            joined = ' '.join(p for p in parts if len(p) > 20)
            if joined:
                return joined

        # 3) fallback: 전체 페이지 텍스트
        full = clean_text(soup.get_text(separator=' '))
        return full[:2000]

    except Exception as e:
        print(f"[!] 본문 추출 오류: {e} – {url}")
        return ""

def clean_text_post(text: str) -> str:
    if pd.isna(text):
        return ""
    text = re.sub(r'©|ⓒ|무단 전재.*?금지|기자\s*[\w\W]*?관련뉴스.*', '', text)
    return re.sub(r'\s+', ' ', text).strip()

def preprocess_news_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    if 'date' in df.columns:
        df['date'] = pd.to_datetime(df['date'], errors='coerce')
        df = df.dropna(subset=['date'])
    if 'content' in df.columns:
        df['content'] = df['content'].apply(clean_text_post)
    return df.drop_duplicates(subset=['title','date'])

def adjust_excel_column_width(filename: str):
    wb = load_workbook(filename)
    ws = wb.active
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 200)
    wb.save(filename)

def get_naver_news_titles(search_query: str, per_month_articles: int, start_date: dt.date, end_date: dt.date) -> pd.DataFrame:
    news_list = []
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
        "Accept-Language": "ko-KR,ko;q=0.9",
    }
    base_url = "https://search.naver.com/search.naver"
    current = dt.date(start_date.year, start_date.month, 1)
    last_month = dt.date(end_date.year, end_date.month, 1)

    while current <= last_month:
        year_month = current.strftime("%Y-%m")
        start_str = current.strftime("%Y.%m.%d")
        temp = (current.replace(day=28) + dt.timedelta(days=4))
        end_of_month = (temp - dt.timedelta(days=temp.day)).strftime("%Y.%m.%d")

        print(f"\n==== {year_month}월 뉴스 검색 중... ====")
        collected, page, seen_urls = 0, 1, set()

        while collected < per_month_articles and page <= 100:
            params = {
                "ssc": "tab.news.all",
                "query": search_query,
                "pd": "3",
                "ds": start_str,
                "de": end_of_month,
                "nso": f"so:r,p:from{start_str.replace('.','')}to{end_of_month.replace('.','')}",
                "start": str((page - 1) * 10 + 1)
            }
            resp = requests.get(base_url, headers=headers, params=params)
            print(f"[디버깅] 페이지: {page}, 응답 코드: {resp.status_code}")
            if resp.status_code != 200:
                break

            soup = BeautifulSoup(resp.text, 'html.parser')
            cards = soup.find_all('div', class_='sds-comps-vertical-layout')
            for card in cards:
                if collected >= per_month_articles:
                    break
                span = card.find('span', class_='sds-comps-text-type-headline1')
                if not span:
                    continue
                a = span.find_parent('a', href=True)
                if not a:
                    continue
                url = a['href']
                if url in seen_urls:
                    continue
                seen_urls.add(url)

                raw_date = None
                for s in card.find_all('span', class_='sds-comps-text-type-body2'):
                    t = s.get_text(strip=True)
                    if re.match(r'^\d{4}\.\d{2}\.\d{2}\.$', t):
                        raw_date = t
                        break
                if not raw_date:
                    continue
                date_str = dt.datetime.strptime(raw_date, '%Y.%m.%d.').strftime('%Y-%m-%d')

                content = get_article_content(url)
                if not content or len(content) < 100:
                    continue

                collected += 1
                news_list.append({
                    "keyword": search_query,
                    "month": year_month,
                    "date": date_str,
                    "title": span.get_text(strip=True),
                    "url": url,
                    "content": content
                })

            page += 1
            time.sleep(random.uniform(2, 5))

        current = (current.replace(day=1) + dt.timedelta(days=32)).replace(day=1)

    return pd.DataFrame(news_list)

if __name__ == "__main__":
    keyword_input = input("[입력] 키워드를 쉼표로 구분하여 입력하세요: ")
    keywords = [kw.strip() for kw in keyword_input.split(',') if kw.strip()]
    per_month_articles = int(input("[입력] 월별 기사 수 (예: 5): "))
    start_input = input("[입력] 뉴스 시작 날짜 (YYYY-MM-DD): ")
    end_input   = input("[입력] 뉴스 끝 날짜 (YYYY-MM-DD): ")
    try:
        sd = dt.datetime.strptime(start_input, "%Y-%m-%d").date()
        ed = dt.datetime.strptime(end_input, "%Y-%m-%d").date()
    except ValueError:
        print("날짜 형식이 잘못되었습니다."); exit(1)
    if sd > ed:
        print("시작일이 종료일보다 늦습니다."); exit(1)

    for kw in keywords:
        df = get_naver_news_titles(kw, per_month_articles, sd, ed)
        if df.empty:
            print(f"[!] {kw} 수집된 뉴스 없음")
            continue
        df = preprocess_news_dataframe(df)
        df['date'] = pd.to_datetime(df['date']).dt.strftime('%Y-%m-%d')
        if 'month' in df.columns:
            df = df.drop(columns=['month'])
        filename = f"naver_newsdata_test_{kw}.xlsx"
        df.to_excel(filename, index=False)
        adjust_excel_column_width(filename)
        print(f"[✓] 저장 완료: {filename}")
